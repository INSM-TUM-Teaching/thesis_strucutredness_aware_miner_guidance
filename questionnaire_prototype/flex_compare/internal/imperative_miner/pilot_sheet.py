from __future__ import annotations

"""
Generate a thesis-friendly imperative miner result report (Markdown first).

The script:
1. loads one XES event log
2. discovers an Inductive Miner process tree
3. creates Petri net and BPMN artifacts
4. computes quantitative evaluation metrics
5. writes a filled Markdown sheet following the target template
6. optionally converts the sheet to PDF (when pandoc is available)
"""

import argparse
import contextlib
import datetime as dt
import fcntl
import getpass
import json
import math
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterator

# openpyxl backs only the auxiliary cross-run xlsx metric aggregate; it is not
# needed for discovery, metrics, or the markdown/Petri/process-tree artifacts the
# flex_compare app actually renders. Keep it optional so a missing openpyxl
# degrades the xlsx export to a no-op instead of crashing every imperative run.
try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    _OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
    DataValidation = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False
import pm4py
from pm4py.objects.log.importer.xes import importer as xes_importer

from flex_compare.internal.experiment_reports import build_report_output_dir
from flex_compare.internal.experiment_reports import infer_experiment_class
from flex_compare.internal.imperative_miner.evaluation import make_report_row
from flex_compare.internal.imperative_miner.evaluation import mine_process_models
from flex_compare.internal.imperative_miner.evaluation import set_log_name
from flex_compare.internal.shared.formatting import checkbox as _checkbox
from flex_compare.internal.shared.formatting import fmt_bool_ja_nein as _fmt_bool_ja_nein
from flex_compare.internal.shared.formatting import fmt_number as _fmt_number_base
from flex_compare.internal.shared.pdf_export import relative_path as _relative_base
from flex_compare.internal.shared.pdf_export import safe_pdf_export as _safe_pdf_export


_EXCEL_SHEET_NAME = "metrics"
_EXCEL_SUMMARY_SHEET_NAME = "summary"
_EXCEL_FILENAME = "imperative_miner_quantitative_metrics.xlsx"
_EXCEL_META_COLUMNS = [
    "log_name",
    "structuredness_class",
    "generated_at",
    "run_id",
    "conformance_method",
    "noise_threshold",
]
_EXCEL_IGNORED_RESULT_KEYS = {"log_id"}
_STRUCTUREDNESS_DROPDOWN_VALUES = ["all", "structured", "semi", "loosely"]


def _fmt_number(value: Any, digits: int = 3) -> str:
    return _fmt_number_base(value, digits=digits, none_placeholder="-")


def _relative(path: Path, base: Path) -> str:
    return _relative_base(path, base)


def _collect_log_stats(log) -> Dict[str, int]:
    n_cases = len(log)
    n_events = sum(len(trace) for trace in log)
    activities = {
        event.get("concept:name")
        for trace in log
        for event in trace
        if event.get("concept:name") is not None
    }
    return {
        "n_cases": n_cases,
        "n_events": n_events,
        "n_activities": len(activities),
    }


def _make_json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _make_json_safe(inner) for key, inner in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_make_json_safe(item) for item in value]
    return str(value)


def _is_excel_numeric_scalar(value: Any) -> bool:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    if isinstance(value, float) and not math.isfinite(value):
        return False
    return True


def _prefixed_numeric_values(
    prefix: str,
    values: Dict[str, Any],
    *,
    ignore_keys: set[str] | None = None,
) -> Dict[str, int | float]:
    numeric_values: Dict[str, int | float] = {}
    ignored = ignore_keys or set()

    for key, value in values.items():
        if key in ignored or not _is_excel_numeric_scalar(value):
            continue
        numeric_values[f"{prefix}{key}"] = value

    return numeric_values


def _build_quantitative_excel_row(
    *,
    log_name: str,
    structuredness_class: str,
    generated_at: str,
    run_id: str,
    conformance_method: str,
    noise_threshold: float,
    log_stats: Dict[str, Any],
    result_align: Dict[str, Any],
    result_token: Dict[str, Any],
) -> Dict[str, Any]:
    row: Dict[str, Any] = {
        "log_name": log_name,
        "structuredness_class": structuredness_class,
        "generated_at": generated_at,
        "run_id": run_id,
        "conformance_method": conformance_method,
        "noise_threshold": noise_threshold,
    }
    row.update(_prefixed_numeric_values("log_", log_stats))
    row.update(
        _prefixed_numeric_values(
            "align_",
            result_align,
            ignore_keys=_EXCEL_IGNORED_RESULT_KEYS,
        )
    )
    row.update(
        _prefixed_numeric_values(
            "token_",
            result_token,
            ignore_keys=_EXCEL_IGNORED_RESULT_KEYS,
        )
    )
    return row


def _ordered_excel_columns(rows: list[Dict[str, Any]]) -> list[str]:
    all_columns = {column for row in rows for column in row.keys()}
    ordered = [column for column in _EXCEL_META_COLUMNS if column in all_columns]
    ordered.extend(sorted(column for column in all_columns if column not in ordered))
    return ordered


def _excel_column_name(index: int) -> str:
    name = ""
    value = index
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _write_metrics_sheet(workbook: Workbook, rows: list[Dict[str, Any]]) -> list[str]:
    if _EXCEL_SHEET_NAME in workbook.sheetnames:
        del workbook[_EXCEL_SHEET_NAME]

    worksheet = workbook.create_sheet(_EXCEL_SHEET_NAME, 0)
    columns = _ordered_excel_columns(rows)
    worksheet.append(columns)

    for row in rows:
        worksheet.append([row.get(column) for column in columns])

    return columns


def _compute_summary_lookup_rows(
    rows: list[Dict[str, Any]],
    metric_names: list[str],
) -> list[Dict[str, Any]]:
    lookup_rows: list[Dict[str, Any]] = []

    for class_name in _STRUCTUREDNESS_DROPDOWN_VALUES:
        if class_name == "all":
            class_rows = rows
        else:
            class_rows = [
                row for row in rows if row.get("structuredness_class") == class_name
            ]

        for metric_name in metric_names:
            values = [
                row.get(metric_name)
                for row in class_rows
                if _is_excel_numeric_scalar(row.get(metric_name))
            ]
            count = len(values)
            lookup_rows.append(
                {
                    "structuredness_class": class_name,
                    "metric": metric_name,
                    "count": count,
                    "mean": sum(values) / count if count else None,
                    "min": min(values) if count else None,
                    "max": max(values) if count else None,
                }
            )

    return lookup_rows


def _summary_lookup_formula(value_col_letter: str, visible_row_idx: int) -> str:
    class_range = f"${_excel_column_name(7)}$2:${_excel_column_name(7)}$1000"
    metric_range = f"${_excel_column_name(8)}$2:${_excel_column_name(8)}$1000"
    value_range = f"${value_col_letter}$2:${value_col_letter}$1000"
    dropdown_ref = "$B$1"
    metric_ref = f"$A{visible_row_idx}"
    count_formula = (
        f"SUMIFS($I$2:$I$1000,{class_range},{dropdown_ref},{metric_range},{metric_ref})"
    )

    if value_col_letter == "I":
        return f"={count_formula}"

    value_formula = (
        f"SUMIFS({value_range},{class_range},{dropdown_ref},{metric_range},{metric_ref})"
    )
    return f'=IF($B{visible_row_idx}=0,"",{value_formula})'


def _write_summary_sheet(workbook: Workbook, columns: list[str], rows: list[Dict[str, Any]]) -> None:
    if _EXCEL_SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[_EXCEL_SUMMARY_SHEET_NAME]

    worksheet = workbook.create_sheet(_EXCEL_SUMMARY_SHEET_NAME)
    worksheet["A1"] = "structuredness_class"
    worksheet["B1"] = "all"
    worksheet["A3"] = "metric"
    worksheet["B3"] = "count"
    worksheet["C3"] = "mean"
    worksheet["D3"] = "min"
    worksheet["E3"] = "max"

    data_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(_STRUCTUREDNESS_DROPDOWN_VALUES)}"',
        allow_blank=False,
    )
    worksheet.add_data_validation(data_validation)
    data_validation.add(worksheet["B1"])

    metric_names = [column for column in columns if column not in _EXCEL_META_COLUMNS]
    lookup_rows = _compute_summary_lookup_rows(rows, metric_names)

    for row_idx, metric_name in enumerate(metric_names, start=4):
        worksheet.cell(row=row_idx, column=1, value=metric_name)
        worksheet.cell(row=row_idx, column=2, value=_summary_lookup_formula("I", row_idx))
        worksheet.cell(row=row_idx, column=3, value=_summary_lookup_formula("J", row_idx))
        worksheet.cell(row=row_idx, column=4, value=_summary_lookup_formula("K", row_idx))
        worksheet.cell(row=row_idx, column=5, value=_summary_lookup_formula("L", row_idx))

    worksheet["G1"] = "structuredness_class"
    worksheet["H1"] = "metric"
    worksheet["I1"] = "count"
    worksheet["J1"] = "mean"
    worksheet["K1"] = "min"
    worksheet["L1"] = "max"

    for row_idx, lookup_row in enumerate(lookup_rows, start=2):
        worksheet.cell(row=row_idx, column=7, value=lookup_row["structuredness_class"])
        worksheet.cell(row=row_idx, column=8, value=lookup_row["metric"])
        worksheet.cell(row=row_idx, column=9, value=lookup_row["count"])
        worksheet.cell(row=row_idx, column=10, value=lookup_row["mean"])
        worksheet.cell(row=row_idx, column=11, value=lookup_row["min"])
        worksheet.cell(row=row_idx, column=12, value=lookup_row["max"])

    worksheet.column_dimensions["G"].hidden = True
    worksheet.column_dimensions["H"].hidden = True
    worksheet.column_dimensions["I"].hidden = True
    worksheet.column_dimensions["J"].hidden = True
    worksheet.column_dimensions["K"].hidden = True
    worksheet.column_dimensions["L"].hidden = True


def _write_excel_rows(excel_path: Path, rows: list[Dict[str, Any]]) -> None:
    workbook = Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    columns = _write_metrics_sheet(workbook, rows)
    _write_summary_sheet(workbook, columns, rows)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)


def _load_excel_rows(excel_path: Path) -> list[Dict[str, Any]]:
    if not excel_path.exists():
        return []

    try:
        workbook = load_workbook(excel_path, data_only=False)
    except (zipfile.BadZipFile, KeyError, OSError) as exc:
        # Quarantine the unreadable file so future runs start fresh instead of
        # being blocked forever. Common cause: a previous run was killed mid-save.
        ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        quarantine = excel_path.with_suffix(f".corrupt-{ts}.xlsx")
        try:
            excel_path.rename(quarantine)
            print(f"[pilot_sheet] corrupt workbook quarantined → {quarantine} ({exc})", flush=True)
        except OSError:
            pass
        return []

    if _EXCEL_SHEET_NAME not in workbook.sheetnames:
        return []

    worksheet = workbook[_EXCEL_SHEET_NAME]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return []

    headers = [str(header) for header in values[0] if header is not None]
    data_rows: list[Dict[str, Any]] = []
    for row_values in values[1:]:
        row_dict = {
            headers[idx]: row_values[idx]
            for idx in range(min(len(headers), len(row_values)))
            if headers[idx] is not None
        }
        if row_dict.get("log_name"):
            data_rows.append(row_dict)

    return data_rows


@contextlib.contextmanager
def _exclusive_file_lock(target: Path) -> Iterator[None]:
    """Serialize read-modify-write on `target` across threads AND processes.

    Uses fcntl.flock on a sibling `.lock` file so the aggregate xlsx can never
    be read by one writer while another is mid-save — the scenario that caused
    the BadZipFile corruption we saw earlier.
    """
    target.parent.mkdir(parents=True, exist_ok=True)
    lock_path = target.with_suffix(target.suffix + ".lock")
    fd = lock_path.open("w")
    try:
        fcntl.flock(fd.fileno(), fcntl.LOCK_EX)
        yield
    finally:
        try:
            fcntl.flock(fd.fileno(), fcntl.LOCK_UN)
        finally:
            fd.close()


def _update_quantitative_metrics_workbook(
    *,
    excel_path: Path,
    row: Dict[str, Any],
) -> None:
    with _exclusive_file_lock(excel_path):
        existing_rows = _load_excel_rows(excel_path)
        updated_rows = [
            existing_row
            for existing_row in existing_rows
            if existing_row.get("log_name") != row["log_name"]
        ]
        updated_rows.append(row)
        updated_rows.sort(key=lambda item: str(item.get("log_name", "")))
        _write_excel_rows(excel_path, updated_rows)


def _safe_visualizations(
    *,
    log,
    output_assets_dir: Path,
    noise_threshold: float,
) -> Dict[str, Any]:
    """
    Try to create process tree, Petri net and BPMN visualizations.

    The function returns generated file paths and error messages (if any),
    without aborting the full sheet generation.
    """
    output_assets_dir.mkdir(parents=True, exist_ok=True)

    paths: Dict[str, Any] = {
        "process_tree_path": None,
        "petri_net_path": None,
        "petri_net_pnml_path": None,
        "bpmn_path": None,
        "model_error": None,
    }

    try:
        process_tree = pm4py.discover_process_tree_inductive(log, noise_threshold=noise_threshold)
        net, im, fm = pm4py.convert_to_petri_net(process_tree)

        process_tree_path = output_assets_dir / "process_tree.png"
        petri_net_path = output_assets_dir / "petri_net.png"
        petri_net_pnml_path = output_assets_dir / "model.pnml"
        bpmn_path = output_assets_dir / "bpmn.png"

        pm4py.save_vis_process_tree(process_tree, str(process_tree_path))
        pm4py.save_vis_petri_net(net, im, fm, str(petri_net_path))

        # PNML export so post-hoc tools (e.g. Entropia entropy-based precision —
        # see miners/shared/entropy_precision.py) can read the discovered net.
        # Visualizations alone are not parseable; the PNML is. Failure here is
        # non-fatal — the rest of the artefact bundle remains usable.
        try:
            pm4py.write_pnml(net, im, fm, str(petri_net_pnml_path))
            paths["petri_net_pnml_path"] = petri_net_pnml_path
        except Exception as pnml_exc:
            paths["petri_net_pnml_path"] = None
            # Append to model_error rather than overwrite; PNML loss is a soft failure.
            existing = paths.get("model_error") or ""
            sep = "; " if existing else ""
            paths["model_error"] = f"{existing}{sep}PNML export failed: {pnml_exc}"

        try:
            bpmn = pm4py.convert_to_bpmn(process_tree)
            pm4py.save_vis_bpmn(bpmn, str(bpmn_path))
            paths["bpmn_path"] = bpmn_path
        except Exception as bpmn_exc:
            paths["bpmn_path"] = None
            paths["model_error"] = f"BPMN visualization failed: {bpmn_exc}"

        paths["process_tree_path"] = process_tree_path
        paths["petri_net_path"] = petri_net_path
    except Exception as exc:
        paths["model_error"] = str(exc)

    return paths


def _build_markdown(
    *,
    run_id: str,
    date_str: str,
    bearbeiter: str,
    log_id: str,
    log_name: str,
    structuredness: str,
    log_path: Path,
    log_stats: Dict[str, int],
    model_ok: bool,
    model_error: str | None,
    model_runtime_sec: float | None,
    process_tree_img: str | None,
    petri_net_img: str | None,
    bpmn_img: str | None,
    result_align: Dict[str, Any],
    result_token: Dict[str, Any],
    report_row: Dict[str, Any],
    output_md_dir: Path,
    pm4py_version: str,
    conformance_method: str,
    noise_threshold: float,
    preprocessing_note: str,
) -> str:
    """
    Render the final report markdown in the requested structure.
    """
    model_saved = process_tree_img is not None or petri_net_img is not None or bpmn_img is not None

    sound_possible = result_align.get("is_wf_net") is True
    sound_status = result_align.get("soundness_status")

    token_fitness_available = result_token.get("fitness_primary") is not None
    align_fitness_available = result_align.get("fitness_primary") is not None
    precision_available = result_align.get("precision") is not None
    f1_available = result_align.get("f1") is not None

    preprocessed = bool(preprocessing_note.strip())
    formatted_date = date_str
    date_iso = dt.datetime.now().isoformat(timespec="seconds")

    discovery_sec = _fmt_number(result_align.get("discovery_runtime_sec"))
    conf_sec = _fmt_number(result_align.get("conformance_runtime_sec"))
    total_sec = _fmt_number(result_align.get("total_runtime_sec"))

    table_line = (
        f"| {report_row.get('Log', '-')} | {report_row.get('Miner', '-')} | "
        f"{_fmt_number(result_align.get('fitness_primary'))} | {_fmt_number(result_align.get('precision'))} | "
        f"{_fmt_number(result_align.get('f1'))} | {_fmt_number(result_align.get('n_places'), 0)} | "
        f"{_fmt_number(result_align.get('n_transitions'), 0)} | {_fmt_number(result_align.get('n_silent_transitions'), 0)} | "
        f"{_fmt_number(result_align.get('n_arcs'), 0)} | {_fmt_bool_ja_nein(result_align.get('is_wf_net'))} | "
        f"{_fmt_bool_ja_nein(result_align.get('is_sound'))} | {discovery_sec} | {conf_sec} | {total_sec} | "
        f"Conformance={conformance_method}, Noise={noise_threshold} |"
    )

    visuals_block = []
    if process_tree_img:
        visuals_block.append(f"- Process Tree: ![Process Tree]({_relative(Path(process_tree_img), output_md_dir)})")
    if petri_net_img:
        visuals_block.append(f"- Petri Net: ![Petri Net]({_relative(Path(petri_net_img), output_md_dir)})")
    if bpmn_img:
        visuals_block.append(f"- BPMN: ![BPMN]({_relative(Path(bpmn_img), output_md_dir)})")
    if not visuals_block:
        visuals_block.append("- No visualization available (error during model visualization).")

    return f"""# Imperative Miner Results Report

## 1. Run Metadata
- Run ID: {run_id}
- Date: {formatted_date}
- Author: {bearbeiter}
- Log ID: {log_id}
- Log name: {log_name}
- Structuredness class: {structuredness}
- Technical timestamp: {date_iso}

---

## 2. Input Log
- File format: {log_path.suffix.lower().lstrip('.') or 'unknown'}
- Number of cases: {log_stats['n_cases']}
- Number of events: {log_stats['n_events']}
- Number of activities: {log_stats['n_activities']}
- Preprocessing performed:
  - {_checkbox(not preprocessed)} no
  - {_checkbox(preprocessed)} yes
- If yes, which?
  - Filtering:
  - Activity renaming:
  - Removing incomplete traces:
  - Other: {preprocessing_note if preprocessed else "-"}
- Observations in the log:
  - [ ] many rare traces
  - [ ] many variants
  - [ ] missing values
  - [ ] unclear labels
  - Notes:

---

## 3. Miner Setup
- Miner: Inductive Miner
- Implementation: PM4Py (`discover_process_tree_inductive` + `convert_to_petri_net`)
- Tool: PM4Py
- Version: {pm4py_version}
- Output format: Process Tree, Petri Net, BPMN (when convertible)
- Parameters:
  - noise threshold: {noise_threshold}
  - variant: default PM4Py inductive setup
  - other: conformance_method={conformance_method}
- Parameter rationale: standard pilot configuration for a reproducible baseline.
- Reproducibly documented:
  - [x] yes
  - [ ] no

---

## 4. Model Generation
- Model successfully generated:
  - {_checkbox(model_ok)} yes
  - {_checkbox(not model_ok)} no
- Runtime (seconds): {_fmt_number(model_runtime_sec)}
- Error message / exception: {model_error or "-"}
- Model saved:
  - {_checkbox(model_saved)} yes
  - {_checkbox(not model_saved)} no
- File path / name: {_relative(output_md_dir, output_md_dir)}
- Model visualization
{chr(10).join(visuals_block)}

---

## 5. Extractable Structural Metrics
# Imperative Discovery Experiment Report

| Log | Miner | Fitness | Precision | F1 | #Places | #Transitions | #Silent | #Arcs | WF-net | Sound | Discovery (s) | Conformance (s) | Total (s) | Notes |
|---|---:|---:|---:|---:|---:|---:|---:|---:|---|---|---:|---:|---:|---|
{table_line}
|  |  |  |  |  |  |  |  |  |  |  |  |  |  |  |
|  |  |  |  |  |  |  |  |  |  |  |  |  |  |  |

- Soundness check technically possible:
  - {_checkbox(sound_possible)} yes
  - {_checkbox(not sound_possible)} no
- Soundness result: {sound_status}
- Other directly extractable structural info:
  - soundness_diagnostics present: {_fmt_bool_ja_nein(result_align.get("soundness_diagnostics") is not None)}
  - n_silent_transitions: {_fmt_number(result_align.get("n_silent_transitions"), 0)}

---

## 6. Extractable Quality Metrics
### Fitness
- token-based replay fitness extractable:
  - {_checkbox(token_fitness_available)} yes
  - {_checkbox(not token_fitness_available)} no
- Value: {_fmt_number(result_token.get("fitness_primary"))}
- alignment fitness extractable:
  - {_checkbox(align_fitness_available)} yes
  - {_checkbox(not align_fitness_available)} no
- Value: {_fmt_number(result_align.get("fitness_primary"))}

### Precision
- precision extractable:
  - {_checkbox(precision_available)} yes
  - {_checkbox(not precision_available)} no
- measure used: {conformance_method}
- Value: {_fmt_number(result_align.get("precision"))}

### Optional
- F-score / F1 extractable:
  - {_checkbox(f1_available)} yes
  - {_checkbox(not f1_available)} no
- Value: {_fmt_number(result_align.get("f1"))}
- Generalization extractable:
  - [ ] yes
  - [x] no
- Value: -

---

## 7. Qualitative initial assessment (fill in later / post-processing)
- Model appears fundamentally plausible:
  - [ ] yes
  - [ ] partly
  - [ ] no
- Observations:
  - [ ] flower-like / over-generalized
  - [ ] very fragmented
  - [ ] many silent transitions
  - [ ] hard to read
  - [ ] unexpected loops
  - [ ] other
- Short note on model structure:
- Short note on readability:
"""


def generate_pilot_sheet(
    *,
    log_path: Path,
    output_root: Path,
    run_id: str | None,
    bearbeiter: str | None,
    noise_threshold: float,
    conformance_method: str,
    preprocessing_note: str,
    export_pdf: bool,
) -> Dict[str, Path | None]:
    """
    Generate one complete result report bundle (markdown + artifacts + optional pdf).
    """
    if not log_path.exists():
        raise FileNotFoundError(f"Log file not found: {log_path}")

    log = xes_importer.apply(str(log_path))
    log_name = log_path.stem
    set_log_name(log, log_name)

    run_id_val = run_id or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    structuredness = infer_experiment_class(log_name, default="custom")
    bearbeiter_val = bearbeiter or getpass.getuser()
    date_str = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    out_dir = build_report_output_dir(output_root, log_name, "inductive_miner")
    assets_dir = out_dir / "assets"
    out_dir.mkdir(parents=True, exist_ok=True)

    log_stats = _collect_log_stats(log)

    model_start = dt.datetime.now()
    visuals = _safe_visualizations(log=log, output_assets_dir=assets_dir, noise_threshold=noise_threshold)
    model_runtime_sec = (dt.datetime.now() - model_start).total_seconds()
    model_ok = visuals.get("process_tree_path") is not None and visuals.get("petri_net_path") is not None

    result_align = mine_process_models(
        [log],
        conformance_method=conformance_method,
        noise_threshold=noise_threshold,
    )[0]
    result_token = mine_process_models(
        [log],
        conformance_method="token_replay",
        noise_threshold=noise_threshold,
    )[0]
    report_row = make_report_row(result_align)
    # The xlsx aggregate is best-effort: without openpyxl the run still yields
    # the markdown, Petri net, process tree and metrics that the app consumes.
    if _OPENPYXL_AVAILABLE:
        excel_path = output_root / _EXCEL_FILENAME
        excel_row = _build_quantitative_excel_row(
            log_name=log_name,
            structuredness_class=structuredness,
            generated_at=date_str,
            run_id=run_id_val,
            conformance_method=conformance_method,
            noise_threshold=noise_threshold,
            log_stats=log_stats,
            result_align=result_align,
            result_token=result_token,
        )
        _update_quantitative_metrics_workbook(excel_path=excel_path, row=excel_row)
    else:
        excel_path = None
        print("[pilot_sheet] openpyxl not installed — skipping xlsx metric "
              "aggregate (run otherwise unaffected).", flush=True)

    markdown = _build_markdown(
        run_id=run_id_val,
        date_str=date_str,
        bearbeiter=bearbeiter_val,
        log_id=log_name,
        log_name=log_name,
        structuredness=structuredness,
        log_path=log_path,
        log_stats=log_stats,
        model_ok=model_ok,
        model_error=visuals.get("model_error"),
        model_runtime_sec=model_runtime_sec,
        process_tree_img=str(visuals["process_tree_path"]) if visuals.get("process_tree_path") else None,
        petri_net_img=str(visuals["petri_net_path"]) if visuals.get("petri_net_path") else None,
        bpmn_img=str(visuals["bpmn_path"]) if visuals.get("bpmn_path") else None,
        result_align=result_align,
        result_token=result_token,
        report_row=report_row,
        output_md_dir=out_dir,
        pm4py_version=pm4py.__version__,
        conformance_method=conformance_method,
        noise_threshold=noise_threshold,
        preprocessing_note=preprocessing_note,
    )

    markdown_path = out_dir / "ergebnisbericht.md"
    markdown_path.write_text(markdown, encoding="utf-8")

    data_path = out_dir / "result_data.json"
    data_payload = {
        "run_id": run_id_val,
        "generated_at": date_str,
        "bearbeiter": bearbeiter_val,
        "log_name": log_name,
        "log_path": str(log_path),
        "structuredness_class": structuredness,
        "log_stats": log_stats,
        "parameters": {
            "noise_threshold": noise_threshold,
            "conformance_method": conformance_method,
            "preprocessing_note": preprocessing_note,
            "pm4py_version": pm4py.__version__,
        },
        "result_align": result_align,
        "result_token": result_token,
        "artifacts": {
            "markdown_path": str(markdown_path),
            "excel_path": str(excel_path) if excel_path else None,
            "process_tree_path": str(visuals["process_tree_path"]) if visuals.get("process_tree_path") else None,
            "petri_net_path": str(visuals["petri_net_path"]) if visuals.get("petri_net_path") else None,
            "petri_net_pnml_path": str(visuals["petri_net_pnml_path"]) if visuals.get("petri_net_pnml_path") else None,
            "bpmn_path": str(visuals["bpmn_path"]) if visuals.get("bpmn_path") else None,
        },
    }
    data_path.write_text(
        json.dumps(_make_json_safe(data_payload), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    pdf_path = _safe_pdf_export(markdown_path) if export_pdf else None

    return {
        "output_dir": out_dir,
        "markdown_path": markdown_path,
        "data_path": data_path,
        "excel_path": excel_path,
        "pdf_path": pdf_path,
        "process_tree_path": visuals.get("process_tree_path"),
        "petri_net_path": visuals.get("petri_net_path"),
        "petri_net_pnml_path": visuals.get("petri_net_pnml_path"),
        "bpmn_path": visuals.get("bpmn_path"),
        "result_align": result_align,
        "result_token": result_token,
        "log_stats": log_stats,
    }


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate an imperative miner result report (Markdown, optional PDF)."
    )
    parser.add_argument(
        "--log-path",
        type=Path,
        required=True,
        help="Path to the input XES log.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path("Experimente"),
        help="Root directory for generated experiment reports.",
    )
    parser.add_argument(
        "--run-id",
        type=str,
        default=None,
        help="Optional manual run identifier.",
    )
    parser.add_argument(
        "--bearbeiter",
        type=str,
        default=None,
        help="Name of the operator/author for the metadata block.",
    )
    parser.add_argument(
        "--noise-threshold",
        type=float,
        default=0.0,
        help="Noise threshold for Inductive Miner discovery.",
    )
    parser.add_argument(
        "--conformance-method",
        type=str,
        default="alignments",
        choices=["alignments", "token_replay"],
        help="Method for the main conformance block in the sheet.",
    )
    parser.add_argument(
        "--preprocessing-note",
        type=str,
        default="",
        help="Optional note describing preprocessing steps.",
    )
    parser.add_argument(
        "--export-pdf",
        action="store_true",
        help="Attempt PDF conversion with pandoc after markdown generation.",
    )
    return parser


def main() -> None:
    parser = _build_arg_parser()
    args = parser.parse_args()

    outputs = generate_pilot_sheet(
        log_path=args.log_path,
        output_root=args.output_root,
        run_id=args.run_id,
        bearbeiter=args.bearbeiter,
        noise_threshold=args.noise_threshold,
        conformance_method=args.conformance_method,
        preprocessing_note=args.preprocessing_note,
        export_pdf=args.export_pdf,
    )

    print(f"Results report generated in: {outputs['output_dir']}")
    print(f"Markdown: {outputs['markdown_path']}")
    print(f"Data: {outputs['data_path']}")
    print(f"Excel: {outputs['excel_path']}")
    if outputs["pdf_path"] is not None:
        print(f"PDF: {outputs['pdf_path']}")
    else:
        print("PDF: not generated")
    print(f"Process Tree image: {outputs['process_tree_path']}")
    print(f"Petri Net image: {outputs['petri_net_path']}")
    print(f"BPMN image: {outputs['bpmn_path']}")


if __name__ == "__main__":
    main()
